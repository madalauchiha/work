#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from xlrd import open_workbook
import sys
from collections import OrderedDict
from common.comm_paras import CASE_XLS_PATH, g_rst_xy, g_mods
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import *


def parse_cell_value(str_cell_value):
    if str_cell_value:
        if 'linux' == sys.platform:
            list_cell_value = str_cell_value.split('\n')

            pair_cell_value = []
            for item in list_cell_value:
                pair_cell_value.append([i.strip().strip(';') for i in item.split(':')])

            return OrderedDict([(item[0], ':'.join(item[1:])) for item in pair_cell_value])
    else:
        return []


def clear_case_result(xls_path, sheet_name):
    wk = load_workbook(xls_path)
    st = wk[sheet_name]

    for rowidx, colidx in g_rst_xy[sheet_name].values():
        cell = st.cell(row=rowidx, column=colidx, value='')

    wk.save(xls_path)


caseno = 0


def get_selected_case_param(tbl_path, sheet_name):
    try:
        startrow, startcol = get_row_col_by_val(tbl_path, sheet_name, 'switch')
    except TypeError as e:
        print('keyword "switch" not found in sheet "%s", please modify case excel!' % sheet_name)
        return []

    wk = open_workbook(tbl_path)
    st = wk.sheet_by_name(sheet_name)

    rows = st.nrows
    datas = []
    g_rst_xy[sheet_name] = dict()

    for i in range(startrow+1, rows):
        switch = st.cell(i, startcol).value
        if (sheet_name in g_mods) and switch == "on":
            global caseno
            caseno += 1

            datas.append([st.cell(i, j).value for j in range(startcol+1, startcol+4)])

            # col_name = startcol + 1
            col_rst = startcol + 4
            # casename = st.cell(i, col_name).value
            # g_rst_xy[sheet_name][casename] = (i+1, col_rst+1)
            g_rst_xy[sheet_name][caseno] = (i + 1, col_rst + 1)

    clear_case_result(tbl_path, sheet_name)

    return datas


def get_row_col_by_val(xls_path, sheet_name, value):
    wk = open_workbook(xls_path)
    st = wk.sheet_by_name(sheet_name)

    rows = st.nrows
    for rowidx in range(rows):
        row = st.row(rowidx)
        for colidx, cell in enumerate(row):
            if value == cell.value:
                return rowidx, colidx


def write_rst_to_xls(xls_path, sheet_name, caseno, rst):
    wk = load_workbook(xls_path)
    st = wk[sheet_name]
    # print(caseno)
    rowidx, colidx = g_rst_xy[sheet_name][caseno]
    cell = st.cell(row=rowidx, column=colidx, value=rst)

    if 'PASS' == rst:
        cell.font = Font(color=BLACK, bold=False)
    else:
        cell.font = Font(color=RED, bold=True)

    wk.save(xls_path)


if __name__ == "__main__":
    print(get_selected_case_param(CASE_XLS_PATH, 'FPGA'))
